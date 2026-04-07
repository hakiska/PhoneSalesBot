"""Каталог товаров — сопоставление коротких названий с полными для 1С."""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

CATALOG_FILE = "catalog.json"

# Сокращения которые используют продавцы
ABBREVIATIONS = {
    "pm": "pro max",
    "ch": "change",
    "g": "glass",
    "or": "original",
    "ori": "original",
    "inc": "incell",
    "sp": "service pack",
    "w/f": "with frame",
    "nf": "no frame",
    "org": "original",
    "oled": "oled",
    "5g": "5g",
    "4g": "4g",
}


def load_catalog() -> dict[str, str]:
    """Загружает сохранённые маппинги short_name -> full_name."""
    path = Path(CATALOG_FILE)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def save_catalog(catalog: dict[str, str]):
    """Сохраняет маппинги."""
    Path(CATALOG_FILE).write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def load_products_from_excel(filepath: str) -> list[str]:
    """Загружает список полных названий из Excel (колонка B)."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    products = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str) and val.strip():
            products.append(val.strip())
    wb.close()
    return products


def load_products_from_json(filepath: str = "products.json") -> list[str]:
    """Загружает список полных названий из JSON-файла."""
    path = Path(filepath)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return []


def _normalize(text: str) -> str:
    """Приводит текст к нижнему регистру, убирает лишние пробелы."""
    return re.sub(r'\s+', ' ', text.lower().strip())


def _expand_abbreviations(words: list[str]) -> list[str]:
    """Раскрывает сокращения: pm -> pro max, ch -> change и т.д."""
    expanded = []
    for w in words:
        if w in ABBREVIATIONS:
            expanded.extend(ABBREVIATIONS[w].split())
        else:
            expanded.append(w)
    return expanded


def _word_score(words: list[str], product_norm: str) -> int:
    """Считает сколько слов из запроса входят в название товара."""
    return sum(1 for w in words if w in product_norm)


def find_product(short_name: str, products: list[str], catalog: dict[str, str]) -> list[str]:
    """Ищет товар по короткому названию.

    1. Проверяет сохранённые маппинги
    2. Раскрывает сокращения (pm -> pro max, ch -> change, etc.)
    3. Ищет по вхождению всех слов
    4. Если не нашли — частичное совпадение

    Returns:
        Список подходящих полных названий.
    """
    key = _normalize(short_name)

    # Точное совпадение в каталоге
    if key in catalog:
        return [catalog[key]]

    words = key.split()

    # Раскрываем сокращения
    expanded = _expand_abbreviations(words)

    # Поиск по вхождению ВСЕХ раскрытых слов
    exact_matches = []
    for product in products:
        norm_product = _normalize(product)
        if all(w in norm_product for w in expanded):
            exact_matches.append(product)

    if len(exact_matches) == 1:
        catalog[key] = exact_matches[0]
        save_catalog(catalog)
        return exact_matches

    if exact_matches:
        return exact_matches

    # Также попробуем с оригинальными словами (без раскрытия)
    orig_matches = []
    for product in products:
        norm_product = _normalize(product)
        if all(w in norm_product for w in words):
            orig_matches.append(product)

    if len(orig_matches) == 1:
        catalog[key] = orig_matches[0]
        save_catalog(catalog)
        return orig_matches

    if orig_matches:
        return orig_matches

    # Частичное совпадение — по раскрытым словам
    min_score = max(1, len(expanded) // 2)
    scored = []
    for product in products:
        norm_product = _normalize(product)
        score = _word_score(expanded, norm_product)
        if score >= min_score:
            scored.append((score, product))

    scored.sort(key=lambda x: x[0], reverse=True)
    partial_matches = [p for _, p in scored[:10]]

    if len(partial_matches) == 1:
        catalog[key] = partial_matches[0]
        save_catalog(catalog)

    return partial_matches


def add_mapping(short_name: str, full_name: str, catalog: dict[str, str]):
    """Вручную добавляет маппинг."""
    key = _normalize(short_name)
    catalog[key] = full_name
    save_catalog(catalog)
