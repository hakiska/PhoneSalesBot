"""Хранение продаж в SQLite и экспорт в Excel."""

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config import DB_PATH


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            seller TEXT NOT NULL,
            product TEXT NOT NULL,
            qty INTEGER NOT NULL,
            price INTEGER NOT NULL,
            total INTEGER NOT NULL,
            payment_type TEXT NOT NULL,
            recipient TEXT DEFAULT '',
            is_debt INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS exchanges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            seller TEXT NOT NULL,
            product_out TEXT NOT NULL,
            price_out INTEGER NOT NULL,
            product_in TEXT NOT NULL,
            price_in INTEGER NOT NULL,
            difference INTEGER NOT NULL,
            payment_type TEXT DEFAULT '',
            recipient TEXT DEFAULT ''
        )
    """)
    # Миграция: добавляем is_debt если его нет
    try:
        conn.execute("SELECT is_debt FROM sales LIMIT 1")
    except sqlite3.OperationalError:
        conn.execute("ALTER TABLE sales ADD COLUMN is_debt INTEGER DEFAULT 0")
    conn.commit()
    return conn


def add_sale(seller: str, product: str, qty: int, price: int,
             total: int, payment_type: str, recipient: str,
             is_debt: bool = False) -> int:
    """Добавляет продажу. Возвращает ID записи."""
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO sales (timestamp, seller, product, qty, price, total, payment_type, recipient, is_debt)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (datetime.now().isoformat(), seller, product, qty, price, total, payment_type, recipient, int(is_debt))
    )
    conn.commit()
    sale_id = cur.lastrowid
    conn.close()
    return sale_id


def delete_last_sale(seller: str) -> bool:
    """Удаляет последнюю продажу продавца. Возвращает True если удалено."""
    conn = get_db()
    row = conn.execute(
        "SELECT id FROM sales WHERE seller = ? ORDER BY id DESC LIMIT 1",
        (seller,)
    ).fetchone()
    if row:
        conn.execute("DELETE FROM sales WHERE id = ?", (row[0],))
        conn.commit()
        conn.close()
        return True
    conn.close()
    return False


def delete_sale_by_id(sale_id: int) -> dict | None:
    """Удаляет продажу по ID. Возвращает удалённую запись или None."""
    conn = get_db()
    row = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
    if row:
        sale = _row_to_dict(row)
        conn.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
        conn.commit()
        conn.close()
        return sale
    conn.close()
    return None


def get_sale_by_id(sale_id: int) -> dict | None:
    """Получает продажу по ID."""
    conn = get_db()
    row = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
    conn.close()
    if row:
        return _row_to_dict(row)
    return None


def partial_return(sale_id: int, return_qty: int) -> dict | None:
    """Частичный возврат — уменьшает кол-во. Если возврат всего — удаляет.
    Возвращает исходную запись или None."""
    conn = get_db()
    row = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
    if not row:
        conn.close()
        return None

    sale = _row_to_dict(row)

    if return_qty >= sale["qty"]:
        # Полный возврат
        conn.execute("DELETE FROM sales WHERE id = ?", (sale_id,))
    else:
        # Частичный — уменьшаем кол-во
        new_qty = sale["qty"] - return_qty
        new_total = new_qty * sale["price"]
        conn.execute(
            "UPDATE sales SET qty = ?, total = ? WHERE id = ?",
            (new_qty, new_total, sale_id)
        )

    conn.commit()
    conn.close()
    return sale


def add_exchange(seller: str, product_out: str, price_out: int,
                 product_in: str, price_in: int,
                 payment_type: str = "", recipient: str = "") -> int:
    """Добавляет обмен. Возвращает ID."""
    conn = get_db()
    difference = price_out - price_in
    cur = conn.execute(
        """INSERT INTO exchanges (timestamp, seller, product_out, price_out, product_in, price_in, difference, payment_type, recipient)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (datetime.now().isoformat(), seller, product_out, price_out, product_in, price_in, difference, payment_type, recipient)
    )
    conn.commit()
    ex_id = cur.lastrowid
    conn.close()
    return ex_id


def get_today_exchanges() -> list[dict]:
    """Обмены за сегодня."""
    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    rows = conn.execute(
        "SELECT * FROM exchanges WHERE timestamp LIKE ? ORDER BY id",
        (f"{today}%",)
    ).fetchall()
    conn.close()
    cols = ["id", "timestamp", "seller", "product_out", "price_out", "product_in", "price_in", "difference", "payment_type", "recipient"]
    return [dict(zip(cols, row)) for row in rows]


def get_today_sales(seller: str = None) -> list[dict]:
    """Продажи за сегодня. Если seller указан — только его."""
    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")

    if seller:
        rows = conn.execute(
            "SELECT * FROM sales WHERE timestamp LIKE ? AND seller = ? ORDER BY id",
            (f"{today}%", seller)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM sales WHERE timestamp LIKE ? ORDER BY id",
            (f"{today}%",)
        ).fetchall()

    conn.close()
    return [_row_to_dict(r) for r in rows]


def get_sales_by_date(date_str: str) -> list[dict]:
    """Продажи за конкретную дату (формат YYYY-MM-DD)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM sales WHERE timestamp LIKE ? ORDER BY id",
        (f"{date_str}%",)
    ).fetchall()
    conn.close()
    return [_row_to_dict(r) for r in rows]


def get_today_debts() -> list[dict]:
    """Долги за сегодня."""
    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    rows = conn.execute(
        "SELECT * FROM sales WHERE timestamp LIKE ? AND is_debt = 1 ORDER BY id",
        (f"{today}%",)
    ).fetchall()
    conn.close()
    return [_row_to_dict(r) for r in rows]


def _row_to_dict(row) -> dict:
    columns = ["id", "timestamp", "seller", "product", "qty", "price", "total", "payment_type", "recipient", "is_debt"]
    d = {}
    for i, col in enumerate(columns):
        if i < len(row):
            d[col] = row[i]
        else:
            d[col] = 0 if col == "is_debt" else ""
    return d


def export_to_excel(sales: list[dict], filename: str = None) -> str:
    """Экспортирует продажи в Excel. Возвращает путь к файлу."""
    if not filename:
        today = datetime.now().strftime("%Y-%m-%d")
        filename = f"sales_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    # Заголовки
    headers = ["No", "Время", "Продавец", "Товар", "Кол-во", "Цена", "Сумма", "Оплата", "Получатель", "Долг"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Данные
    debt_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for i, sale in enumerate(sales, 1):
        ts = datetime.fromisoformat(sale["timestamp"])
        row_data = [
            i,
            ts.strftime("%H:%M"),
            sale["seller"],
            sale["product"],
            sale["qty"],
            sale["price"],
            sale["total"],
            sale["payment_type"],
            sale["recipient"],
            "Долг" if sale.get("is_debt") else "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = thin_border
            if col in (5, 6, 7):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            if sale.get("is_debt"):
                cell.fill = debt_fill

    # Итоговая строка
    total_row = len(sales) + 2
    ws.cell(row=total_row, column=6, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(s["total"] for s in sales)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = "#,##0"

    # Итоги по типам оплаты
    summary_row = total_row + 2
    ws.cell(row=summary_row, column=1, value="Сводка по оплате:").font = Font(bold=True, size=11)

    cash_total = sum(s["total"] for s in sales if s["payment_type"] == "Наличные" and not s.get("is_debt"))
    debt_total = sum(s["total"] for s in sales if s.get("is_debt"))

    ws.cell(row=summary_row + 1, column=1, value="Наличные:")
    ws.cell(row=summary_row + 1, column=2, value=cash_total).number_format = "#,##0"

    # Разбивка по получателям Каспи
    recipients = {}
    for s in sales:
        if s["payment_type"] == "Каспи" and s["recipient"] and not s.get("is_debt"):
            recipients[s["recipient"]] = recipients.get(s["recipient"], 0) + s["total"]

    row = summary_row + 2
    for name, amount in sorted(recipients.items()):
        ws.cell(row=row, column=1, value=f"Каспи ({name}):")
        ws.cell(row=row, column=2, value=amount).number_format = "#,##0"
        row += 1

    if debt_total:
        row += 1
        ws.cell(row=row, column=1, value="В долг:").font = Font(bold=True, color="FF0000")
        ws.cell(row=row, column=2, value=debt_total).number_format = "#,##0"

    # Обмены
    exchanges = get_today_exchanges()
    if exchanges:
        row += 2
        ws.cell(row=row, column=1, value="Обмены:").font = Font(bold=True, size=11)
        row += 1
        for ex in exchanges:
            ts = datetime.fromisoformat(ex["timestamp"])
            ws.cell(row=row, column=1, value=ts.strftime("%H:%M"))
            ws.cell(row=row, column=2, value=f"{ex['product_out']} ({ex['price_out']:,})")
            ws.cell(row=row, column=3, value="->")
            ws.cell(row=row, column=4, value=f"{ex['product_in']} ({ex['price_in']:,})")
            diff = ex["difference"]
            label = f"Возврат клиенту: {diff:,}" if diff > 0 else f"Доплата: {abs(diff):,}"
            ws.cell(row=row, column=5, value=label)
            row += 1

    # Ширина колонок
    widths = [5, 8, 12, 30, 8, 12, 12, 12, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filepath = str(Path(filename).resolve())
    wb.save(filepath)
    return filepath
